import math, io
from datetime import datetime, timedelta, date
from copy import copy as xcopy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.cell import MergedCell

TIER_COLORS = {"High":"FF1F3864","Medium":"FFFF8C00","Low":"FFFFC000","Minimal":"FF92D050"}
FONT_COLORS = {"High":"FFFFFFFF","Medium":"FF006400","Low":"FF006400","Minimal":"FF006400"}

def apply_tier(cell, t):
    if t not in TIER_COLORS: t = "Minimal"
    cell.fill = PatternFill("solid", start_color=TIER_COLORS[t], end_color=TIER_COLORS[t])
    cell.font = Font(bold=True, color=FONT_COLORS[t], name="Calibri", size=10)

def copy_row_fmt(src_ws, src_row, dst_ws, dst_row, max_col):
    for col in range(2, max_col + 1):
        src = src_ws.cell(row=src_row, column=col)
        dst = dst_ws.cell(row=dst_row, column=col)
        if isinstance(dst, MergedCell): continue
        if src.has_style:
            dst.font = xcopy(src.font); dst.fill = xcopy(src.fill)
            dst.border = xcopy(src.border); dst.alignment = xcopy(src.alignment)
            dst.number_format = src.number_format

def safe_set(ws, row, col, val):
    c = ws.cell(row=row, column=col)
    if not isinstance(c, MergedCell): c.value = val

def excel_round(val, decimals=0):
    factor = 10 ** decimals
    return math.floor(val * factor + 0.5) / factor

def calc_score(insider, float_pct, ev_sales, ev_ebitda, d, e, f=0):
    A = excel_round(min(insider / max(float_pct, 0.01), 5) / 5 * 30, 1)
    B = excel_round(min(insider / 100, 1) * 25, 1)
    # Valuation Risk (0-20): no earnings = max risk (20). NM means risk, not zero.
    def _is_nm(x):
        try:
            float(x); return False
        except (ValueError, TypeError):
            return True
    _s_nm = _is_nm(ev_sales)
    _e_nm = _is_nm(ev_ebitda)
    if _s_nm and _e_nm:
        cS, cE = 20.0, 0.0
    else:
        cS = 10.0 if _s_nm else min(float(ev_sales) / 5 * 9, 9)
        cE = 10.0 if _e_nm else min(max(float(ev_ebitda) - 5, 0) / 25 * 9, 9)
    C   = excel_round(cS + cE, 1)
    F   = excel_round(max(-30, min(30, float(f))), 1)
    raw = excel_round(A + B + C + d + e + F, 1)
    score = int(excel_round(max(0, min(100, raw)), 0))
    tier  = "High" if score>=75 else "Medium" if score>=50 else "Low" if score>=25 else "Minimal"
    return score, tier, A, B, C, raw

def score_from_sm_row(wsSM, row):
    try:
        e=wsSM.cell(row=row,column=5).value; f=wsSM.cell(row=row,column=6).value
        i=wsSM.cell(row=row,column=9).value; j=wsSM.cell(row=row,column=10).value
        l=wsSM.cell(row=row,column=12).value; m=wsSM.cell(row=row,column=13).value
        n=wsSM.cell(row=row,column=17).value  # F score (col Q)
        if e is None: return 0
        # Float (col 6) is a derived formula (=100-insider); compute directly
        try:
            fval = float(f)
        except (ValueError, TypeError):
            fval = round(100 - float(e), 2)
        try:
            fscore = float(n)
        except (ValueError, TypeError):
            fscore = 0
        score, _, _, _, _, _ = calc_score(float(e), fval, i, j, float(l), float(m), fscore)
        return score
    except: return 0

def add_name_to_workbook(file_bytes, params):
    wb = load_workbook(io.BytesIO(file_bytes), keep_vba=True)
    wsSS = wb["Short Screen"]; wsSM = wb["Scoring Model"]
    wsLV = wb["Lockup Verification"]; wsHB = wb["Historical Backtest"]

    TICKER       = params["ticker"]
    COMPANY      = params["company"]
    prosp_date   = datetime.strptime(params["prospectus_date"], "%Y-%m-%d").date()
    LOCKUP_DAYS  = int(params["lockup_days"])
    SEC_SOURCE   = params["sec_source"]
    SPONSOR      = params["sponsor"]
    INSIDER_PCT  = float(params["insider_pct"])
    EV_SALES     = str(params["ev_sales"])
    EV_EBITDA    = str(params["ev_ebitda"])
    D_SCORE      = float(params["d_score"])
    MODIFIER     = float(params["modifier"])
    EARLY_RELEASE = params.get("early_release", "No")

    lockup_date = prosp_date + timedelta(days=LOCKUP_DAYS)
    float_pct   = round(100 - INSIDER_PCT, 2)

    nextRow = 5
    while wsSS.cell(row=nextRow, column=3).value: nextRow += 1

    sm_cell = wsSM.cell(row=nextRow, column=3)
    if isinstance(sm_cell, MergedCell) or sm_cell.value:
        raise ValueError(f"SM row {nextRow} not empty — SS/SM out of sync.")

    nextRowLV = 5
    while wsLV.cell(row=nextRowLV, column=2).value: nextRowLV += 1
    nextRowHB = 12
    while wsHB.cell(row=nextRowHB, column=2).value: nextRowHB += 1

    score, tier, A, B, C, raw = calc_score(INSIDER_PCT, float_pct, EV_SALES, EV_EBITDA, D_SCORE, MODIFIER)

    days_out = (lockup_date - date.today()).days
    date_str = f"EXPIRED {abs(days_out)}d ago" if days_out < 0 else f"{days_out} days out"
    val_str  = "Pre-revenue / NM valuation" if EV_SALES == "NM" else f"EV/Sales {EV_SALES}x"
    if EV_EBITDA != "NM": val_str += f", EV/EBITDA {EV_EBITDA}x"
    d_map = {25:"Active Form 4-confirmed sellers.",22:"Multiple confirmed Form 4 sellers.",
             20:"Early lockup release triggered.",15:"VC/PE overhang upcoming — no confirmed selling yet.",
             12:"Lockup expired; selling evidence present.",8:"PE/VC upcoming; no confirmed selling.",
             5:"Corporate parent / minimal mechanism.",0:"No supply shock mechanism."}
    d_str = next((v for k,v in sorted(d_map.items(),reverse=True) if D_SCORE >= k), "No mechanism.")
    m_str = ("Anti-short: insider buying or strong institutional demand offsets overhang." if MODIFIER <= -3
             else "Mild anti-short evidence." if MODIFIER < 0
             else "Confirmed selling beyond D-score." if MODIFIER >= 3
             else "Additional selling signals." if MODIFIER > 0 else "")
    thesis = f"{SPONSOR} ({INSIDER_PCT}%) lockup {lockup_date.strftime('%b %d %Y')} — {date_str}.  {val_str}.  {d_str}  {m_str}".strip()

    # Write SM
    for merge in list(wsSM.merged_cells.ranges):
        if merge.min_row <= nextRow <= merge.max_row: wsSM.unmerge_cells(str(merge))
    copy_row_fmt(wsSM, 6, wsSM, nextRow, 16)
    safe_set(wsSM, nextRow, 2, "=ROW()-4"); safe_set(wsSM, nextRow, 3, TICKER)
    safe_set(wsSM, nextRow, 4, COMPANY);   safe_set(wsSM, nextRow, 5, INSIDER_PCT)
    safe_set(wsSM, nextRow, 6, round(100 - float(INSIDER_PCT), 2))  # hardcoded float (matches existing rows)
    safe_set(wsSM, nextRow, 7, f"=IFERROR(ROUND(MIN(E{nextRow}/MAX(F{nextRow},0.01),5)/5*30,1),0)")
    safe_set(wsSM, nextRow, 8, f"=ROUND(MIN(E{nextRow}/100,1)*25,1)")
    safe_set(wsSM, nextRow, 9, EV_SALES); safe_set(wsSM, nextRow, 10, EV_EBITDA)
    safe_set(wsSM, nextRow, 11, f'=ROUND(IF(AND(ISERROR(VALUE(I{nextRow})),ISERROR(VALUE(J{nextRow}))),20,IF(ISERROR(VALUE(I{nextRow})),10,MIN(VALUE(I{nextRow})/5*9,9))+IF(ISERROR(VALUE(J{nextRow})),10,MIN(MAX(VALUE(J{nextRow})-5,0)/25*9,9))),1)')
    safe_set(wsSM, nextRow, 12, D_SCORE); safe_set(wsSM, nextRow, 13, MODIFIER)
    safe_set(wsSM, nextRow, 14, f"=ROUND(G{nextRow}+H{nextRow}+K{nextRow}+L{nextRow}+M{nextRow}+Q{nextRow},1)")  # Raw (col N) includes F from Q
    safe_set(wsSM, nextRow, 15, f"=MAX(0,MIN(100,ROUND(N{nextRow},0)))")  # Final Score (col O)
    safe_set(wsSM, nextRow, 16, f'=IF(O{nextRow}>=75,"High",IF(O{nextRow}>=50,"Medium",IF(O{nextRow}>=25,"Low","Minimal")))')  # Tier (col P)
    safe_set(wsSM, nextRow, 17, 0)  # F score placeholder (col Q) — updated daily by momentum updater
    apply_tier(wsSM.cell(row=nextRow, column=15), tier)
    apply_tier(wsSM.cell(row=nextRow, column=16), tier)

    # Write SS
    copy_row_fmt(wsSS, 6, wsSS, nextRow, 14)
    # Clear row background — only score/tier cols get tier color; row fill depends on tier
    TIER_ROW_BG = {"High":"FF1F3864","Medium":"FFFFEB9C","Low":"FFFFFFFF","Minimal":"FFFFFFFF"}
    row_bg = TIER_ROW_BG.get(tier, "FFFFFFFF")
    for col in [2,3,4,5,6,7,8,9,10,11,14]:
        c = wsSS.cell(row=nextRow, column=col)
        if not isinstance(c, MergedCell):
            c.fill = PatternFill("solid", start_color=row_bg, end_color=row_bg)
    safe_set(wsSS, nextRow, 2, "=ROW()-4"); safe_set(wsSS, nextRow, 3, TICKER)
    safe_set(wsSS, nextRow, 4, COMPANY);   safe_set(wsSS, nextRow, 5, lockup_date)
    safe_set(wsSS, nextRow, 6, f"=E{nextRow}-TODAY()"); safe_set(wsSS, nextRow, 7, prosp_date)
    safe_set(wsSS, nextRow, 8, INSIDER_PCT); safe_set(wsSS, nextRow, 9, round(100 - float(INSIDER_PCT), 2))  # hardcoded float
    safe_set(wsSS, nextRow, 10, EV_SALES);  safe_set(wsSS, nextRow, 11, EV_EBITDA)
    safe_set(wsSS, nextRow, 12, f"='Scoring Model'!O{nextRow}")
    safe_set(wsSS, nextRow, 13, f"='Scoring Model'!P{nextRow}")
    safe_set(wsSS, nextRow, 14, thesis)
    apply_tier(wsSS.cell(row=nextRow, column=12), tier)
    apply_tier(wsSS.cell(row=nextRow, column=13), tier)

    # Sort SS and SM together
    last_row = nextRow
    rows_ss = []
    for r in range(5, last_row + 1):
        row_data = {c: wsSS.cell(row=r,column=c).value for c in range(2,15)}
        row_fmt  = {c: wsSS.cell(row=r,column=c).number_format for c in range(2,15)}
        row_fill = {c: xcopy(wsSS.cell(row=r,column=c).fill) for c in range(2,15)}
        row_font = {c: xcopy(wsSS.cell(row=r,column=c).font) for c in range(2,15)}
        rows_ss.append({'data':row_data,'fmt':row_fmt,'fill':row_fill,'font':row_font,
                        'score':score_from_sm_row(wsSM,r),'ticker':row_data.get(3)})
    rows_sm = []
    for r in range(5, last_row + 1):
        row_data={}; row_fmt={}; row_fill={}; row_font={}
        for c in range(2,18):
            cell = wsSM.cell(row=r,column=c)
            if isinstance(cell,MergedCell): row_data[c]=None; continue
            row_data[c]=cell.value; row_fmt[c]=cell.number_format
            row_fill[c]=xcopy(cell.fill); row_font[c]=xcopy(cell.font)
        rows_sm.append({'data':row_data,'fmt':row_fmt,'fill':row_fill,'font':row_font,'ticker':row_data.get(3)})

    rows_ss.sort(key=lambda x: x['score'], reverse=True)
    ticker_order = [r['ticker'] for r in rows_ss]
    sm_by_ticker = {r['ticker']: r for r in rows_sm}
    rows_sm_sorted = [sm_by_ticker[t] for t in ticker_order if t in sm_by_ticker]

    TIER_ROW_BG = {"High":"FF1F3864","Medium":"FFFFEB9C","Low":"FFFFFFFF","Minimal":"FFFFFFFF"}
    for idx, row in enumerate(rows_ss):
        r = 5 + idx
        for c in range(2,15):
            cell = wsSS.cell(row=r,column=c)
            cell.value=row['data'][c]; cell.number_format=row['fmt'][c]
            cell.fill=row['fill'][c]; cell.font=row['font'][c]
        wsSS.cell(row=r,column=2).value=f"=ROW()-4"
        wsSS.cell(row=r,column=6).value=f"=E{r}-TODAY()"
        wsSS.cell(row=r,column=12).value=f"='Scoring Model'!O{r}"
        wsSS.cell(row=r,column=13).value=f"='Scoring Model'!P{r}"
        s=row['score']; t="High" if s>=75 else "Medium" if s>=50 else "Low" if s>=25 else "Minimal"
        apply_tier(wsSS.cell(row=r,column=12),t); apply_tier(wsSS.cell(row=r,column=13),t)
        # Re-apply correct row background after sort restores old fills
        row_bg = TIER_ROW_BG.get(t, "FFFFFFFF")
        for col in [2,3,4,5,6,7,8,9,10,11,14]:
            c2 = wsSS.cell(row=r, column=col)
            if not isinstance(c2, MergedCell):
                c2.fill = PatternFill("solid", start_color=row_bg, end_color=row_bg)

    for idx, row in enumerate(rows_sm_sorted):
        r = 5 + idx
        for c in range(2,18):
            cell=wsSM.cell(row=r,column=c)
            if isinstance(cell,MergedCell): continue
            cell.value=row['data'].get(c)
            if c in row['fmt']: cell.number_format=row['fmt'][c]
            if c in row['fill']: cell.fill=row['fill'][c]
            if c in row['font']: cell.font=row['font'][c]
        wsSM.cell(row=r,column=2).value=f"=ROW()-4"
        wsSM.cell(row=r,column=7).value=f"=IFERROR(ROUND(MIN(E{r}/MAX(F{r},0.01),5)/5*30,1),0)"
        wsSM.cell(row=r,column=8).value=f"=ROUND(MIN(E{r}/100,1)*25,1)"
        wsSM.cell(row=r,column=11).value=f'=ROUND(IF(AND(ISERROR(VALUE(I{r})),ISERROR(VALUE(J{r}))),20,IF(ISERROR(VALUE(I{r})),10,MIN(VALUE(I{r})/5*9,9))+IF(ISERROR(VALUE(J{r})),10,MIN(MAX(VALUE(J{r})-5,0)/25*9,9))),1)'
        wsSM.cell(row=r,column=14).value=f"=ROUND(G{r}+H{r}+K{r}+L{r}+M{r}+Q{r},1)"
        wsSM.cell(row=r,column=15).value=f"=MAX(0,MIN(100,ROUND(N{r},0)))"
        wsSM.cell(row=r,column=16).value=f'=IF(O{r}>=75,"High",IF(O{r}>=50,"Medium",IF(O{r}>=25,"Low","Minimal")))'
        s=rows_ss[idx]['score']; t="High" if s>=75 else "Medium" if s>=50 else "Low" if s>=25 else "Minimal"
        apply_tier(wsSM.cell(row=r,column=15),t); apply_tier(wsSM.cell(row=r,column=16),t)
        # Fix SM row background
        sm_bg = TIER_ROW_BG.get(t, "FFFFFFFF")
        for col in list(range(2, 15)) + [17]:
            c2 = wsSM.cell(row=r, column=col)
            if not isinstance(c2, MergedCell):
                c2.fill = PatternFill("solid", start_color=sm_bg, end_color=sm_bg)

    # Ensure SM has a blank buffer row after last data row before rubric
    last_data_row = 5 + len(rows_sm_sorted) - 1
    buffer_row = last_data_row + 1
    buf_cell = wsSM.cell(row=buffer_row, column=3)
    if not isinstance(buf_cell, MergedCell) and buf_cell.value and not str(buf_cell.value).startswith('='):
        wsSM.insert_rows(buffer_row)

    # Write LV
    copy_row_fmt(wsLV,5,wsLV,nextRowLV,12)
    safe_set(wsLV,nextRowLV,2,TICKER); safe_set(wsLV,nextRowLV,3,COMPANY)
    safe_set(wsLV,nextRowLV,4,prosp_date); safe_set(wsLV,nextRowLV,5,LOCKUP_DAYS)
    safe_set(wsLV,nextRowLV,6,lockup_date)
    safe_set(wsLV,nextRowLV,7,f'=IF(F{nextRowLV}<TODAY(),"EXPIRED — "&TEXT(TODAY()-F{nextRowLV},"0")&"d ago","VERIFIED — DATE CONFIRMED")')
    safe_set(wsLV,nextRowLV,8,SEC_SOURCE); safe_set(wsLV,nextRowLV,9,EARLY_RELEASE)
    safe_set(wsLV,nextRowLV,10,"No"); safe_set(wsLV,nextRowLV,11,"No"); safe_set(wsLV,nextRowLV,12,"")

    # Write HB
    copy_row_fmt(wsHB,12,wsHB,nextRowHB,6)
    safe_set(wsHB,nextRowHB,2,TICKER); safe_set(wsHB,nextRowHB,3,COMPANY)
    safe_set(wsHB,nextRowHB,4,lockup_date); safe_set(wsHB,nextRowHB,5,SPONSOR)
    safe_set(wsHB,nextRowHB,6,float_pct)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.read(), score, tier, days_out
