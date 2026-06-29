"""
Momentum updater — appends live 1W/1M price + Form 4 activity
to the thesis cell (col N) for each ticker on Short Screen.
Zero Anthropic API usage. Free data only: Yahoo Finance + SEC EDGAR.
"""
import io, time, logging, re, requests
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

log = logging.getLogger(__name__)

HEADERS = {"User-Agent": "Brandon Ross brandonr1010@gmail.com"}
MOMENTUM_TAG = "| Momentum update:"  # tag so we can find/replace the appended line

def get_cik(ticker):
    """Look up CIK from SEC EDGAR company tickers JSON."""
    try:
        r = requests.get(
            "https://data.sec.gov/submissions/company_tickers.json",
            headers=HEADERS, timeout=15
        )
        if r.status_code != 200:
            return None
        data = r.json()
        for k, v in data.items():
            if v.get('ticker', '').upper() == ticker.upper():
                return str(v['cik_str']).zfill(10)
        return None
    except:
        return None

def get_form4_activity(cik, days_back=30):
    """
    Check recent Form 4 filings for a CIK.
    Returns a short string like "2 insider sales" or "No Form 4s filed"
    """
    try:
        url = f"https://data.sec.gov/submissions/CIK{cik}.json"
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        recent = data.get('filings', {}).get('recent', {})
        forms = recent.get('form', [])
        dates = recent.get('filedAt', recent.get('filingDate', []))
        cutoff = (date.today() - timedelta(days=days_back)).isoformat()
        form4s = [dates[i] for i in range(len(forms)) if forms[i] == '4' and dates[i] >= cutoff]
        if not form4s:
            return f"No Form 4s past {days_back}d"
        return f"{len(form4s)} Form 4 filing{'s' if len(form4s)>1 else ''} past {days_back}d"
    except:
        return None

def get_price_momentum(ticker):
    """Returns (1D_pct, 1W_pct, 1M_pct) via Yahoo Finance."""
    try:
        import yfinance as yf
        hist = yf.Ticker(ticker).history(period="35d")
        if hist.empty or len(hist) < 2:
            return None, None, None
        cur = hist['Close'].iloc[-1]
        d1 = round((cur - hist['Close'].iloc[-2]) / hist['Close'].iloc[-2] * 100, 1) if len(hist) >= 2 else None
        w1 = round((cur - hist['Close'].iloc[-6]) / hist['Close'].iloc[-6] * 100, 1) if len(hist) >= 6 else None
        m1 = round((cur - hist['Close'].iloc[-22]) / hist['Close'].iloc[-22] * 100, 1) if len(hist) >= 22 else None
        return d1, w1, m1
    except:
        return None, None, None

def fmt_pct(val):
    if val is None: return "N/A"
    return f"+{val}%" if val >= 0 else f"{val}%"

def strip_old_momentum(thesis):
    """Remove previously appended momentum line."""
    if not thesis:
        return thesis
    idx = thesis.find(MOMENTUM_TAG)
    if idx != -1:
        return thesis[:idx].rstrip()
    return thesis

def update_momentum(file_bytes, newsapi_key=None):
    """
    For each ticker in Short Screen:
    1. Pull 1W/1M price from Yahoo Finance
    2. Pull recent Form 4 count from EDGAR
    3. Pull sector ETF momentum + news sentiment (F score)
    4. Append/replace momentum line in thesis cell (col N)
    Returns updated file bytes.
    """
    import os
    if not newsapi_key:
        newsapi_key = os.environ.get("NEWSAPI_KEY")

    wb = load_workbook(io.BytesIO(file_bytes), keep_vba=True)
    wsSS = wb["Short Screen"]

    # Get all tickers first
    tickers = []
    for r in range(5, 200):
        t = wsSS.cell(row=r, column=3).value
        if not t: break
        tickers.append(t)

    # Cache CIK lookup
    cik_map = {}
    try:
        r = requests.get(
            "https://data.sec.gov/submissions/company_tickers.json",
            headers=HEADERS, timeout=15
        )
        if r.status_code == 200:
            data = r.json()
            for k, v in data.items():
                cik_map[v.get('ticker','').upper()] = str(v['cik_str']).zfill(10)
            log.info(f"CIK map loaded: {len(cik_map)} tickers")
        time.sleep(0.3)
    except Exception as e:
        log.warning(f"CIK batch load failed: {e}")

    # Get sector context (ETF + news per sector, cached). Stock momentum added per-ticker.
    sector_context = {}
    try:
        from sector import get_sector_context
        sector_context = get_sector_context(tickers, newsapi_key)
        log.info(f"Sector context computed for {len(sector_context)} sectors")
    except Exception as e:
        log.error(f"Sector context FAILED: {type(e).__name__}: {e}")
        import traceback
        log.error(traceback.format_exc())
        sector_context = {}

    updated = 0
    movers_data = []
    for row in range(5, 200):
        ticker = wsSS.cell(row=row, column=3).value
        if not ticker: break

        log.info(f"  Updating momentum for {ticker}")

        # Price momentum
        d1, w1, m1 = get_price_momentum(ticker)
        time.sleep(0.3)

        # Form 4 activity
        f4_str = None
        cik = cik_map.get(ticker.upper())
        if cik:
            f4_str = get_form4_activity(cik, days_back=30)
            time.sleep(0.35)

        # Compute F using stock momentum (w1/m1) + cached sector context
        try:
            from sector import score_ticker
            f_score, sector, _bd = score_ticker(ticker, w1, m1, sector_context)
        except Exception as e:
            log.warning(f"  {ticker}: score_ticker failed: {e}")
            f_score, sector = 0, "Unknown"

        # Collect for Daily Movers tab (1-day move since last scan)
        company = wsSS.cell(row=row, column=4).value or ""
        movers_data.append({
            "ticker": ticker, "company": company, "d1": d1,
            "w1": w1, "m1": m1, "sector": sector,
        })

        # Build momentum line
        price_str = f"1D {fmt_pct(d1)} / 1W {fmt_pct(w1)} / 1M {fmt_pct(m1)}"
        f4_part = f" | {f4_str}" if f4_str else ""
        sector_part = f" | Sector ({sector}): F={'+' if f_score>=0 else ''}{f_score}"
        momentum_line = f"{MOMENTUM_TAG} {price_str}{f4_part}{sector_part}"

        # Write F score to SM col 17 (Q) — feeds into Raw formula
        sm_f_cell = wb["Scoring Model"].cell(row=row, column=14)
        if not isinstance(sm_f_cell, MergedCell):
            sm_f_cell.value = f_score
            sm_f_cell.number_format = '+0;-0;0'
            from openpyxl.styles import Alignment as Aln
            sm_f_cell.alignment = Aln(horizontal='center')

        # Update thesis cell
        thesis_cell = wsSS.cell(row=row, column=14)
        if isinstance(thesis_cell, MergedCell):
            continue
        old_thesis = str(thesis_cell.value) if thesis_cell.value else ""
        clean_thesis = strip_old_momentum(old_thesis)
        thesis_cell.value = f"{clean_thesis}  {momentum_line}" if clean_thesis else momentum_line
        updated += 1
        log.info(f"    {ticker}: {momentum_line}")

    log.info(f"Momentum updated for {updated} tickers — re-sorting by updated scores...")

    # Re-sort SS and SM after F scores are written
    try:
        _resort_after_f_update(wb)
        log.info("Re-sort complete.")
    except Exception as e:
        log.error(f"Re-sort failed: {e}")

    # Build/refresh Daily Movers tab
    try:
        _build_daily_movers(wb, movers_data)
        log.info("Daily Movers tab updated.")
    except Exception as e:
        log.error(f"Daily Movers build failed: {e}")
        import traceback
        log.error(traceback.format_exc())

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _resort_after_f_update(wb):
    """Re-sort Short Screen and Scoring Model by updated scores after F is written."""
    from copy import copy as xcopy
    from openpyxl.styles import PatternFill
    from openpyxl.cell.cell import MergedCell as MC

    wsSS = wb["Short Screen"]
    wsSM = wb["Scoring Model"]

    TIER_COLORS = {"High":"FF1F3864","Medium":"FFFF8C00","Low":"FFFFC000","Minimal":"FF92D050"}
    TIER_ROW_BG = {"High":"FF1F3864","Medium":"FFFFEB9C","Low":"FFFFFFFF","Minimal":"FFFFFFFF"}
    FONT_COLORS = {"High":"FFFFFFFF","Medium":"FF006400","Low":"FF006400","Minimal":"FF006400"}

    def get_score_from_inputs(r):
        """Compute score directly from SM inputs including F."""
        import math
        def er(val, dec=0):
            f = 10**dec; return math.floor(val*f+0.5)/f
        try:
            ins = float(wsSM.cell(row=r,column=5).value or 0)
            # Float is a derived formula (=100-insider); compute it directly
            _fv = wsSM.cell(row=r,column=6).value
            try:
                flt = float(_fv)
            except (ValueError, TypeError):
                flt = round(100 - ins, 2)
            ev_s = wsSM.cell(row=r,column=9).value
            ev_e = wsSM.cell(row=r,column=10).value
            d = float(wsSM.cell(row=r,column=12).value or 0)
            e = float(wsSM.cell(row=r,column=13).value or 0)
            f = float(wsSM.cell(row=r,column=14).value or 0)
            A = er(min(ins/max(flt,0.01),5)/5*30,1)
            B = er(min(ins/100,1)*25,1)
            try: cS=min(float(ev_s)/5*10,10)
            except: cS=0
            try: cE=min(max(float(ev_e)-5,0)/25*10,10)
            except: cE=0
            C = er(cS+cE,1)
            F = er(max(-30,min(30,f)),1)
            raw = er(A+B+C+d+e+F,1)
            score = int(er(max(0,min(100,raw)),0))
            tier = "High" if score>=75 else "Medium" if score>=50 else "Low" if score>=25 else "Minimal"
            return score, tier
        except: return 0, "Minimal"

    # Collect all rows
    rows_ss = []; rows_sm = []
    last_row = 5
    while wsSS.cell(row=last_row,column=3).value: last_row+=1
    last_row -= 1

    for r in range(5, last_row+1):
        ss_data={c:wsSS.cell(row=r,column=c).value for c in range(2,15)}
        ss_fmt={c:wsSS.cell(row=r,column=c).number_format for c in range(2,15)}
        ss_fill={c:xcopy(wsSS.cell(row=r,column=c).fill) for c in range(2,15)}
        ss_font={c:xcopy(wsSS.cell(row=r,column=c).font) for c in range(2,15)}
        score,tier=get_score_from_inputs(r)
        rows_ss.append({'data':ss_data,'fmt':ss_fmt,'fill':ss_fill,'font':ss_font,'score':score,'tier':tier,'ticker':ss_data.get(3)})

        sm_data={c:(None if isinstance(wsSM.cell(row=r,column=c),MC) else wsSM.cell(row=r,column=c).value) for c in range(2,18)}
        sm_fmt={c:(wsSM.cell(row=r,column=c).number_format if not isinstance(wsSM.cell(row=r,column=c),MC) else '') for c in range(2,18)}
        sm_fill={c:(xcopy(wsSM.cell(row=r,column=c).fill) if not isinstance(wsSM.cell(row=r,column=c),MC) else None) for c in range(2,18)}
        sm_font={c:(xcopy(wsSM.cell(row=r,column=c).font) if not isinstance(wsSM.cell(row=r,column=c),MC) else None) for c in range(2,18)}
        rows_sm.append({'data':sm_data,'fmt':sm_fmt,'fill':sm_fill,'font':sm_font,'ticker':sm_data.get(3)})

    rows_ss.sort(key=lambda x:x['score'],reverse=True)
    ticker_order=[r['ticker'] for r in rows_ss]
    sm_by_ticker={r['ticker']:r for r in rows_sm}
    rows_sm_sorted=[sm_by_ticker[t] for t in ticker_order if t in sm_by_ticker]

    from openpyxl.styles import Font as Fnt
    for idx,row in enumerate(rows_ss):
        r=5+idx
        for c in range(2,15):
            cell=wsSS.cell(row=r,column=c)
            cell.value=row['data'][c]; cell.number_format=row['fmt'][c]
            cell.fill=row['fill'][c]; cell.font=row['font'][c]
        wsSS.cell(row=r,column=2).value=f"=ROW()-4"
        wsSS.cell(row=r,column=6).value=f"=E{r}-TODAY()"
        wsSS.cell(row=r,column=12).value=f"='Scoring Model'!P{r}"
        wsSS.cell(row=r,column=13).value=f"='Scoring Model'!Q{r}"
        t=row['tier']
        sc=TIER_COLORS[t]; bg=TIER_ROW_BG[t]; fc=FONT_COLORS[t]
        for col in [2,3,4,5,6,7,8,9,10,11,14]:
            c2=wsSS.cell(row=r,column=col)
            if not isinstance(c2,MC): c2.fill=PatternFill("solid",start_color=bg,end_color=bg)
        for col in [12,13]:
            c2=wsSS.cell(row=r,column=col)
            if not isinstance(c2,MC):
                c2.fill=PatternFill("solid",start_color=sc,end_color=sc)
                c2.font=Fnt(bold=True,color=fc,name="Calibri",size=10)

    for idx,row in enumerate(rows_sm_sorted):
        r=5+idx
        for c in range(2,18):
            cell=wsSM.cell(row=r,column=c)
            if isinstance(cell,MC): continue
            cell.value=row['data'].get(c)
            if row['fmt'].get(c): cell.number_format=row['fmt'][c]
            if row['fill'].get(c): cell.fill=row['fill'][c]
            if row['font'].get(c): cell.font=row['font'][c]
        wsSM.cell(row=r,column=2).value=f"=ROW()-4"
        wsSM.cell(row=r,column=7).value=f"=IFERROR(ROUND(MIN(E{r}/MAX(F{r},0.01),5)/5*30,1),0)"
        wsSM.cell(row=r,column=8).value=f"=ROUND(MIN(E{r}/100,1)*25,1)"
        wsSM.cell(row=r,column=11).value=f"=ROUND(IFERROR(MIN(IFERROR(VALUE(I{r}),0)/5*10,10),0)+IFERROR(MIN(MAX(IFERROR(VALUE(J{r}),0)-5,0)/25*10,10),0),1)"
        wsSM.cell(row=r,column=15).value=f"=ROUND(G{r}+H{r}+K{r}+L{r}+M{r}+N{r},1)"
        wsSM.cell(row=r,column=16).value=f"=MAX(0,MIN(100,ROUND(O{r},0)))"
        wsSM.cell(row=r,column=17).value=f'=IF(P{r}>=75,"High",IF(P{r}>=50,"Medium",IF(P{r}>=25,"Low","Minimal")))'
        s,t=get_score_from_inputs(r)
        t_str="High" if s>=75 else "Medium" if s>=50 else "Low" if s>=25 else "Minimal"
        sc=TIER_COLORS[t_str]; bg=TIER_ROW_BG[t_str]; fc=FONT_COLORS[t_str]
        from openpyxl.styles import Font as Fnt2
        for col in range(2,16):
            c2=wsSM.cell(row=r,column=col)
            if not isinstance(c2,MC): c2.fill=PatternFill("solid",start_color=bg,end_color=bg)
        for col in [16,17]:
            c2=wsSM.cell(row=r,column=col)
            if not isinstance(c2,MC):
                c2.fill=PatternFill("solid",start_color=sc,end_color=sc)
                c2.font=Fnt2(bold=True,color=fc,name="Calibri",size=10)

    # Ensure buffer row
    last_data=5+len(rows_sm_sorted)-1
    buf=last_data+1
    buf_cell=wsSM.cell(row=buf,column=3)
    if not isinstance(buf_cell,MC) and buf_cell.value and not str(buf_cell.value).startswith('='):
        wsSM.insert_rows(buf)


def _build_daily_movers(wb, movers_data):
    """
    Build/refresh the 'Daily Movers' tab: top 10 stocks by absolute 1-day move
    (since last scan), both directions, ranked largest to smallest.
    Rebuilt fresh each scan.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    NAVY = "FF1F3864"; WHITE = "FFFFFFFF"; LIGHT = "FFF0F4FA"
    GREEN = "FF006400"; RED = "FFC00000"; GRAY = "FF808080"
    UP_BG = "FFE2EFDA"; DOWN_BG = "FFFCE4E4"

    SHEET = "Daily Movers"
    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
        # Clear existing content
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
    else:
        ws = wb.create_sheet(SHEET)
        # Position right after Short Screen
        try:
            idx = wb.sheetnames.index("Short Screen") + 1
            wb.move_sheet(SHEET, offset=idx - wb.sheetnames.index(SHEET))
        except Exception:
            pass

    # Filter to names with a valid 1-day move, de-dupe by ticker (keep largest abs move)
    valid = [m for m in movers_data if m.get("d1") is not None]
    by_ticker = {}
    for m in valid:
        tk = m["ticker"]
        if tk not in by_ticker or abs(m["d1"]) > abs(by_ticker[tk]["d1"]):
            by_ticker[tk] = m
    valid = list(by_ticker.values())
    valid.sort(key=lambda m: abs(m["d1"]), reverse=True)
    top = valid[:10]

    thin = Side(style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("B2:G2")
    t = ws["B2"]
    t.value = f"DAILY MOVERS — Biggest 1-Day Moves (updated {datetime.now().strftime('%m/%d/%Y')})"
    t.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
    t.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Subtitle
    ws.merge_cells("B3:G3")
    s = ws["B3"]
    s.value = "Ranked by absolute 1-day price move since last scan. Green = up, Red = down."
    s.fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)
    s.font = Font(italic=True, color="FF333333", size=8, name="Calibri")
    s.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["Rank", "Ticker", "Company", "1-Day", "1-Week", "1-Month", "Sector"]
    cols = [2, 3, 4, 5, 6, 7, 8]
    hdr_row = 5
    for col, h in zip(cols, headers):
        c = ws.cell(row=hdr_row, column=col)
        c.value = h
        c.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
        c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # Data rows
    for i, m in enumerate(top):
        r = hdr_row + 1 + i
        d1 = m["d1"]
        up = d1 >= 0
        row_bg = UP_BG if up else DOWN_BG
        move_color = GREEN if up else RED

        vals = [
            (2, i + 1, "center", None),
            (3, m["ticker"], "center", None),
            (4, m["company"], "left", None),
            (5, d1 / 100, "center", move_color),
            (6, (m["w1"] / 100 if m["w1"] is not None else None), "center", None),
            (7, (m["m1"] / 100 if m["m1"] is not None else None), "center", None),
            (8, m["sector"], "center", None),
        ]
        for col, val, align, color in vals:
            c = ws.cell(row=r, column=col)
            c.value = val
            c.fill = PatternFill("solid", start_color=row_bg, end_color=row_bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = border
            if col in (5, 6, 7):
                c.number_format = "+0.0%;-0.0%;0.0%"
            if col == 3:
                c.font = Font(bold=True, color="FF000000", size=10, name="Calibri")
            elif col == 5:
                c.font = Font(bold=True, color=color, size=10, name="Calibri")
            else:
                c.font = Font(color="FF000000", size=9, name="Calibri")

    # If no data
    if not top:
        ws.merge_cells(f"B{hdr_row+1}:H{hdr_row+1}")
        c = ws.cell(row=hdr_row+1, column=2)
        c.value = "No price data available this scan."
        c.font = Font(italic=True, color=GRAY, size=9)
        c.alignment = Alignment(horizontal="center")

    # Column widths
    widths = {"A": 2, "B": 6, "C": 9, "D": 30, "E": 10, "F": 10, "G": 10, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.sheet_view.showGridLines = False
    return ws
