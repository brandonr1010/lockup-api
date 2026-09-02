import os, io, json, math, time, requests, logging, re
from datetime import datetime, timedelta, date
from supabase import create_client

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

SUPABASE_URL  = os.environ["SUPABASE_URL"]
SUPABASE_KEY  = os.environ["SUPABASE_SERVICE_KEY"]
ANTHROPIC_KEY = os.environ["ANTHROPIC_API_KEY"]
BUCKET        = "workbook"
FILE_NAME     = "Lockup_automation2.xlsm"
MAX_PER_RUN   = 5

SEC_HEADERS = {"User-Agent": "Brandon Ross brandonr1010@gmail.com"}

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def excel_round(val, decimals=0):
    factor = 10 ** decimals
    return math.floor(val * factor + 0.5) / factor

def calc_score(insider, float_pct, ev_sales, ev_ebitda, d, e):
    A = excel_round(min(insider / max(float_pct, 0.01), 5) / 5 * 30, 1)
    B = excel_round(min(insider / 100, 1) * 25, 1)
    # Valuation Risk (0-20): no earnings = max risk (20). NM means risk, not zero.
    def _is_nm(x):
        try:
            float(x); return False
        except (ValueError, TypeError):
            return True
    _s_nm = _is_nm(ev_sales)
    _e_nm = _is_nm(ev_ebitda)
    if _s_nm and _e_nm:
        cS, cE = 20.0, 0.0
    else:
        cS = 10.0 if _s_nm else min(float(ev_sales) / 5 * 9, 9)
        cE = 10.0 if _e_nm else min(max(float(ev_ebitda) - 5, 0) / 25 * 9, 9)
    C   = excel_round(cS + cE, 1)
    raw = excel_round(A + B + C + d + e, 1)
    score = int(excel_round(max(0, min(100, raw)), 0))
    tier  = "High" if score>=75 else "Medium" if score>=50 else "Low" if score>=25 else "Minimal"
    return score, tier

def download_workbook():
    return supabase.storage.from_(BUCKET).download(FILE_NAME)

def upload_workbook(file_bytes):
    supabase.storage.from_(BUCKET).update(
        FILE_NAME, file_bytes,
        {"content-type": "application/vnd.ms-excel.sheet.macroEnabled.12", "upsert": "true"}
    )

def fetch_recent_filings(days_back=7):
    today  = date.today()
    qtr    = (today.month - 1) // 3 + 1
    cutoff = today - timedelta(days=days_back)
    url    = f"https://www.sec.gov/Archives/edgar/full-index/{today.year}/QTR{qtr}/company.idx"

    log.info(f"Fetching EDGAR index: {url}")
    time.sleep(0.5)

    try:
        r = requests.get(url, headers=SEC_HEADERS, timeout=60)
        log.info(f"Status: {r.status_code}, Size: {len(r.content)} bytes")
        if r.status_code != 200:
            log.error(f"Failed: {r.text[:200]}")
            return []
    except Exception as e:
        log.error(f"Request error: {e}")
        return []

    filings = []
    seen = set()
    lines = r.text.split('\n')
    log.info(f"Total lines: {len(lines)}")

    for line in lines:
        if '424B4' not in line:
            continue
        try:
            parts = re.split(r'\s{2,}', line.strip())
            if len(parts) < 4:
                continue
            company   = parts[0].strip()
            form_type = parts[1].strip()
            filed     = parts[3].strip()
            if form_type != '424B4':
                continue
            if not company or company in seen:
                continue
            filed_date = date.fromisoformat(filed[:10])
            if filed_date < cutoff:
                continue
            seen.add(company)
            filings.append({"entity_name": company, "file_date": filed[:10]})
            log.info(f"  Found: {company} ({filed[:10]})")
        except:
            continue

    log.info(f"424B4 filings in last {days_back} days: {len(filings)}")
    return filings

def research_ticker(company_name, filing_date):
    prompt = f"""Research the company "{company_name}" which filed a 424B4 prospectus on {filing_date}.
Return ONLY a JSON object, no other text:
{{
  "ticker": "exchange ticker symbol or empty string",
  "company": "full legal company name",
  "prospectus_date": "YYYY-MM-DD",
  "lockup_days": 180,
  "sec_source": "SEC 424B4 {filing_date}",
  "sponsor": "primary PE/VC sponsor or insider with ownership %",
  "insider_pct": 0.0,
  "ev_sales": "number (compute: Enterprise Value / trailing revenue). Only 'NM' if truly pre-revenue.",
  "ev_ebitda": "number (compute: Enterprise Value / trailing EBITDA). Only 'NM' if EBITDA is negative/zero.",
  "d_score": 0,
  "modifier": 0,
  "early_release": "No",
  "skip": false,
  "skip_reason": ""
}}
Set skip=true if: debt offering, shelf offering, warrant, SPAC, ETF, mutual fund, no lockup.
D-score: 25=active Form4 sellers,22=multiple Form4,20=early release,15=VC/PE upcoming,
12=lockup expired selling,8=PE/VC upcoming,5=corporate parent,0=no mechanism.
Modifier -5 to +5. insider_pct=% held by insiders. Numbers not strings for numeric fields."""

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=1000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}]
        )
        text = "".join(b.text for b in response.content if b.type == "text")
        s = text.find("{"); e = text.rfind("}")
        if s == -1: return None
        parsed = json.loads(text[s:e+1])
        log.info(f"  -> ticker={parsed.get('ticker')} skip={parsed.get('skip')}")
        return parsed
    except Exception as e:
        log.error(f"Research error: {e}")
        return None

def log_entry(params, score, tier):
    try:
        supabase.table("lockup_entries").insert({
            "ticker": params["ticker"], "company": params["company"],
            "prospectus_date": params["prospectus_date"], "lockup_days": int(params["lockup_days"]),
            "insider_pct": float(params["insider_pct"]), "ev_sales": str(params["ev_sales"]),
            "ev_ebitda": str(params["ev_ebitda"]), "d_score": float(params["d_score"]),
            "modifier": float(params["modifier"]), "score": score, "tier": tier,
            "sec_source": params.get("sec_source",""), "sponsor": params.get("sponsor",""),
            "early_release": params.get("early_release","No"),
        }).execute()
        log.info(f"Logged {params['ticker']}")
    except Exception as e:
        log.error(f"Supabase log error: {e}")

def get_workbook_tickers(file_bytes):
    """Tickers currently in the workbook (Scoring Model = source of truth)."""
    from openpyxl import load_workbook as _lwb
    from openpyxl.cell.cell import MergedCell as _MC
    wb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
    ws = wb["Scoring Model"]
    out = set()
    for r in range(5, 300):
        c = ws.cell(row=r, column=3)
        if isinstance(c, _MC):
            continue
        if not c.value:
            break
        out.add(str(c.value).upper().strip())
    return out

def reconcile_table_to_workbook(file_bytes):
    """Backfill ONLY explicitly listed names that were logged to lockup_entries
    but never landed in the workbook. The table is append-only history — it also
    holds names deliberately deleted from the workbook (old duplicates, skips),
    so a blanket table-vs-workbook diff would re-add junk. Newest row per ticker
    wins. Returns (file_bytes, n_backfilled)."""
    BACKFILL_TICKERS = {"BLSM", "ATTO", "APMD", "CISS", "AGCC"}
    wb_tickers = get_workbook_tickers(file_bytes)
    # benched names stay benched — re-adding them just makes enforcement
    # re-bench every run (the CISS/AGCC churn)
    benched = set()
    try:
        from openpyxl import load_workbook as _lwb
        _wbb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
        if "_bench" in _wbb.sheetnames:
            _b = _wbb["_bench"]
            benched = {str(_b.cell(row=r, column=1).value).upper().strip()
                       for r in range(2, 300) if _b.cell(row=r, column=1).value}
    except Exception as e:
        log.error(f"bench read failed: {e}")
    todo = BACKFILL_TICKERS - wb_tickers - benched
    if not todo:
        return file_bytes, 0
    res = supabase.table("lockup_entries").select("*") \
        .order("created_at", desc=True).execute()
    latest = {}
    for row in (res.data or []):
        t = str(row.get("ticker", "")).upper().strip()
        if t in todo and t not in latest:
            latest[t] = row
    log.info(f"Reconcile: backfilling {sorted(latest.keys())}")
    from lockup_engine import add_name_to_workbook
    n = 0
    for t, row in latest.items():
        params = {
            "ticker": row["ticker"], "company": row["company"],
            "prospectus_date": row["prospectus_date"], "lockup_days": row["lockup_days"],
            "insider_pct": row["insider_pct"], "ev_sales": row["ev_sales"],
            "ev_ebitda": row["ev_ebitda"], "d_score": row["d_score"],
            "modifier": row["modifier"], "sec_source": row.get("sec_source", ""),
            "sponsor": row.get("sponsor", ""), "early_release": row.get("early_release", "No"),
        }
        try:
            file_bytes, score, tier, days_out = add_name_to_workbook(file_bytes, params)
            n += 1
            log.info(f"  Backfilled {row['ticker']} (score {score}, {tier})")
        except Exception as e:
            log.error(f"  Backfill failed for {row['ticker']}: {e}")
    return file_bytes, n

def sync_historical_backtest(file_bytes):
    """Rebuild Historical Backtest Section 2 from the Short Screen every run:
    one contiguous row per ticker, all formulas self-row referenced. Prevents
    the row-drift / missing-formula corruption from recurring."""
    from openpyxl import load_workbook as _lwb
    from openpyxl.styles import Font, Alignment
    from copy import copy

    wb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
    hb, ss = wb["Historical Backtest"], wb["Short Screen"]

    # kill any stray merges inside the data area (caused the PTRN/KARD corruption)
    for m in [m for m in list(hb.merged_cells.ranges) if m.min_row >= 12]:
        hb.unmerge_cells(str(m))

    # preserve existing static data (sponsor text is richest here)
    existing = {}
    for r in range(12, 200):
        t = hb.cell(row=r, column=2).value
        if not t:
            continue
        t = str(t).strip()
        if t.startswith(("NOTE", "SECTION")):
            continue
        existing[t] = {"company": hb.cell(row=r, column=3).value,
                       "date": hb.cell(row=r, column=4).value,
                       "sponsor": hb.cell(row=r, column=5).value,
                       "float": hb.cell(row=r, column=6).value}

    # sponsor fallback from lockup_entries
    sponsors = {}
    try:
        res = supabase.table("lockup_entries").select("ticker,sponsor").execute()
        sponsors = {str(x["ticker"]).upper(): x.get("sponsor") for x in (res.data or [])}
    except Exception as e:
        log.error(f"Sponsor lookup failed: {e}")

    universe = []
    for r in range(5, 300):
        t = ss.cell(row=r, column=3).value
        if not t:
            break
        universe.append({"ticker": str(t).strip(),
                         "company": ss.cell(row=r, column=4).value,
                         "lockup": ss.cell(row=r, column=5).value,
                         "float": ss.cell(row=r, column=9).value,
                         "thesis": ss.cell(row=r, column=14).value or ""})

    tpl  = {c: copy(hb.cell(row=12, column=c)._style) for c in range(2, 18)}
    tfmt = {c: hb.cell(row=12, column=c).number_format for c in range(2, 18)}

    for r in range(12, 12 + max(len(universe), len(existing)) + 20):
        for c in range(2, 18):
            hb.cell(row=r, column=c).value = None

    r = 12
    for u in universe:
        ex = existing.get(u["ticker"], {})
        sp = (ex.get("sponsor") or sponsors.get(u["ticker"].upper())
              or u["thesis"][:80].strip())
        vals = {2: u["ticker"], 3: ex.get("company") or u["company"],
                4: ex.get("date") or u["lockup"], 5: sp,
                6: ex.get("float") if ex.get("float") is not None else u["float"],
                7: f'=IFERROR(_xlfn.SINGLE(_xlfn.STOCKHISTORY("{u["ticker"]}",N{r},N{r},0,0,1)),"")',
                8: f'=IFERROR(_xlfn.SINGLE(_xlfn.STOCKHISTORY("{u["ticker"]}",O{r},O{r},0,0,1)),"")',
                9: f'=IFERROR(_xlfn.SINGLE(_xlfn.STOCKHISTORY("{u["ticker"]}",P{r},P{r},0,0,1)),"")',
                10: f'=IFERROR(_xlfn.SINGLE(_xlfn.STOCKHISTORY("{u["ticker"]}",Q{r},Q{r},0,0,1)),"")',
                11: f'=IFERROR(H{r}/G{r}-1,"")', 12: f'=IFERROR(I{r}/G{r}-1,"")',
                13: f'=IFERROR(J{r}/G{r}-1,"")',
                14: f'=WORKDAY(D{r},-1)', 15: f'=WORKDAY(D{r},1)',
                16: f'=WORKDAY(D{r},5)', 17: f'=WORKDAY(D{r},10)'}
        for c, v in vals.items():
            cell = hb.cell(row=r, column=c)
            cell.value = v
            cell._style = copy(tpl[c])
            cell.number_format = tfmt[c]
        r += 1

    nc = hb.cell(row=r + 1, column=2)
    nc.value = ("NOTE: STOCKHISTORY() requires Excel 365 with an active market data feed. "
                "Price cells auto-populate once the lockup date passes; % Chg columns compute "
                "from T-1 close. Rows sync automatically from the Short Screen on every scan.")
    nc.font = Font(italic=True, color="FF808080", size=9, name="Calibri")
    nc.alignment = Alignment(horizontal="left")

    out = io.BytesIO()
    wb.save(out)
    log.info(f"Historical Backtest synced: {len(universe)} rows")
    return out.getvalue()

LIQ_GATE_MM = 5.0  # $MM 30-day avg dollar volume; matches Scoring Model header "≥$5MM"

def resolve_liquidity_gate(file_bytes):
    """Force every Scoring Model liquidity gate (col S) to PASS or FAIL — never
    'pending'. If ADV (col R) is blank, retry via yfinance on whatever trading
    days exist; if still no data, FAIL (unverifiable liquidity = untradeable
    for a short screen)."""
    from openpyxl import load_workbook as _lwb
    wb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
    ws = wb["Scoring Model"]
    changed = 0
    for r in range(5, 300):
        t = ws.cell(row=r, column=3).value
        if not t:
            break
        adv = ws.cell(row=r, column=18).value  # R
        if adv is None or str(adv).strip() == "":
            adv = _fetch_adv_mm(str(t).strip())
            if adv is not None:
                ws.cell(row=r, column=18).value = round(adv, 1)
        gate_cell = ws.cell(row=r, column=19)  # S
        try:
            gate = "PASS" if float(adv) >= LIQ_GATE_MM else "FAIL"
        except (TypeError, ValueError):
            gate = "FAIL"  # no data obtainable
        if gate_cell.value != gate:
            gate_cell.value = gate
            changed += 1
            log.info(f"  Liq gate {t}: ADV={adv} -> {gate}")
    out = io.BytesIO()
    wb.save(out)
    log.info(f"Liquidity gate resolved: {changed} cells updated")
    return out.getvalue()

_px_cache = {}

def _get_history(ticker, days=60):
    """One throttled yfinance history per ticker per run (Yahoo rate limits —
    the Aug 27 run returned nan for every ticker after call volume spiked)."""
    t = str(ticker).upper().strip()
    if t in _px_cache:
        return _px_cache[t]
    try:
        import yfinance as yf
        time.sleep(0.4)
        h = yf.Ticker(t).history(period=f"{days}d")
    except Exception as e:
        log.error(f"  history fetch failed {t}: {e}")
        h = None
    _px_cache[t] = h
    return h

def _fetch_adv_mm(ticker):
    """30-day (or available-days) average daily dollar volume in $MM, or None."""
    try:
        h = _get_history(ticker)
        if h is None or h.empty:
            return None
        dv = (h["Close"] * h["Volume"]).dropna()
        if len(dv) == 0:
            return None
        return float(dv.tail(30).mean()) / 1e6
    except Exception as e:
        log.error(f"  ADV fetch failed for {ticker}: {e}")
        return None

# ── Universe rules: exchange + market cap ────────────────────────────────────
ALLOWED_EXCHANGES = {
    "NYQ": "NYSE", "NMS": "Nasdaq", "NGM": "Nasdaq", "NCM": "Nasdaq",
    "TOR": "TSX",
}
MIN_MCAP_USD = 500e6  # $500M floor — names below are benched, not shown

_profile_cache = {}

def _get_profile(ticker):
    """(exchange_code, market_cap_usd) via yfinance; (None, None) if unavailable."""
    t = str(ticker).upper().strip()
    if t in _profile_cache:
        return _profile_cache[t]
    exch, mcap = None, None
    try:
        import yfinance as yf
        info = yf.Ticker(t).info or {}
        exch = info.get("exchange")
        mcap = info.get("marketCap")
        if mcap is None:
            p, so = info.get("regularMarketPrice"), info.get("sharesOutstanding")
            if p and so:
                mcap = p * so
    except Exception as e:
        log.error(f"  Profile fetch failed for {t}: {e}")
    _profile_cache[t] = (exch, mcap)
    return exch, mcap

def qualifies(ticker):
    """(bool, reason). Fails on non-NYSE/Nasdaq/TSX exchange, mcap < $500M,
    or missing data (benched until data appears — new IPOs re-check daily)."""
    exch, mcap = _get_profile(ticker)
    if exch is None and mcap is None:
        return False, "no exchange/mcap data yet"
    if exch not in ALLOWED_EXCHANGES:
        return False, f"exchange {exch} not NYSE/Nasdaq/TSX"
    if mcap is None:
        return False, "no market cap data yet"
    if mcap < MIN_MCAP_USD:
        return False, f"mcap ${mcap/1e6:,.0f}M < $500M"
    return True, f"{ALLOWED_EXCHANGES[exch]} ${mcap/1e9:,.1f}B"

import re as _re

def _rewrite_row_refs(formula, new_row):
    """Rewrite every single-cell row reference in a self-row formula to new_row."""
    return _re.sub(r"([A-Z]{1,2})\d+(?![0-9])", lambda m: f"{m.group(1)}{new_row}", formula)

def enforce_universe_rules(file_bytes):
    """Remove non-qualifying and duplicate tickers from all display sheets,
    park them on hidden _bench, and promote benched names that now qualify
    (re-added from lockup_entries via lockup_engine). Rewrites self-row
    formulas and ranks after row deletion. Returns file_bytes."""
    from openpyxl import load_workbook as _lwb
    wb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
    ss, sm = wb["Short Screen"], wb["Scoring Model"]

    # current universe from Scoring Model
    tickers, seen = [], set()
    dupes = set()
    for r in range(5, 300):
        t = sm.cell(row=r, column=3).value
        if not t:
            break
        t = str(t).upper().strip()
        if t in seen:
            dupes.add(t)
        seen.add(t)
        tickers.append(t)

    verdicts = {t: qualifies(t) for t in set(tickers)}
    drop = {t for t, (ok, _) in verdicts.items() if not ok}
    for t in sorted(set(tickers)):
        ok, reason = verdicts[t]
        log.info(f"  Universe {t}: {'KEEP' if ok else 'BENCH'} ({reason})")
    if dupes:
        log.info(f"  Duplicates to collapse: {sorted(dupes)}")

    # ---- bench sheet
    if "_bench" not in wb.sheetnames:
        b = wb.create_sheet("_bench")
        b.sheet_state = "hidden"
        b["A1"], b["B1"], b["C1"] = "ticker", "benched_at", "reason"
    bench = wb["_bench"]
    benched_now = {str(bench.cell(row=r, column=1).value).upper().strip()
                   for r in range(2, 300) if bench.cell(row=r, column=1).value}

    # ---- delete rows for dropped/duplicate tickers, sheet by sheet
    def _is_ticker(v):
        s = str(v or "").strip()
        return 1 <= len(s) <= 6 and s == s.upper() and " " not in s and s.isascii() and any(ch.isalpha() for ch in s)

    def purge(ws, tick_col, start=5):
        # find last TICKER row (rubric/doc text below the table must not count)
        # so blank spacer rows mid-table don't truncate the scan
        last = 0
        for rr in range(start, ws.max_row + 1):
            if _is_ticker(ws.cell(row=rr, column=tick_col).value):
                last = rr
        removed, kept_seen = 0, set()
        r = start
        while r <= last:
            t = ws.cell(row=r, column=tick_col).value
            if not _is_ticker(t):
                ws.delete_rows(r, 1)   # compact blank spacer
                last -= 1
                removed += 1
                continue
            t = str(t).upper().strip()
            if t in drop or t in kept_seen:
                ws.delete_rows(r, 1)
                last -= 1
                removed += 1
                continue
            kept_seen.add(t)
            r += 1
        return removed

    purge(ss, 3); purge(sm, 3)
    if "Lockup Verification" in wb.sheetnames:
        purge(wb["Lockup Verification"], 2)
    if "Sources" in wb.sheetnames:
        srcs = wb["Sources"]
        r = 5
        while r <= srcs.max_row:
            t = srcs.cell(row=r, column=2).value
            if t and str(t).upper().strip() in drop:
                srcs.delete_rows(r, 1)
                continue
            r += 1

    # ---- rewrite EVERY formula cell's row refs + ranks (openpyxl does not
    #      adjust formula text on delete_rows; O/P and any future formula
    #      columns must shift too — fixed columns caused the v44 score bug)
    def fix_rows(ws, tick_col, rank=True):
        for r in range(5, 300):
            if not _is_ticker(ws.cell(row=r, column=tick_col).value):
                break
            if rank:
                ws.cell(row=r, column=2).value = r - 4
            for c in range(2, ws.max_column + 1):
                f = ws.cell(row=r, column=c).value
                if isinstance(f, str) and f.startswith("="):
                    ws.cell(row=r, column=c).value = _rewrite_row_refs(f, r)
    fix_rows(ss, 3)
    fix_rows(sm, 3)
    if "Lockup Verification" in wb.sheetnames:
        fix_rows(wb["Lockup Verification"], 2, rank=False)
    # ---- purge benched tickers from _movers_state (stale rows create ghost movers)
    if "_movers_state" in wb.sheetnames:
        ms = wb["_movers_state"]
        r = 2
        while r <= ms.max_row:
            t = ms.cell(row=r, column=1).value
            if t and str(t).upper().strip() in drop:
                ms.delete_rows(r, 1)
                continue
            r += 1

    # ---- record newly benched
    next_b = 2
    while bench.cell(row=next_b, column=1).value:
        next_b += 1
    for t in sorted(drop - benched_now):
        bench.cell(row=next_b, column=1).value = t
        bench.cell(row=next_b, column=2).value = date.today().isoformat()
        bench.cell(row=next_b, column=3).value = verdicts[t][1]
        next_b += 1

    out = io.BytesIO()
    wb.save(out)
    file_bytes = out.getvalue()
    log.info(f"Universe enforced: {len(drop)} benched, dupes collapsed: {sorted(dupes) or 'none'}")

    # ---- promotion: benched names that now qualify, with data in lockup_entries
    wb_tickers = get_workbook_tickers(file_bytes)
    promote = [t for t in benched_now
               if t not in wb_tickers and t not in drop and qualifies(t)[0]]
    if promote:
        try:
            res = supabase.table("lockup_entries").select("*")                 .order("created_at", desc=True).execute()
            latest = {}
            for row in (res.data or []):
                t = str(row.get("ticker", "")).upper().strip()
                if t in promote and t not in latest:
                    latest[t] = row
            from lockup_engine import add_name_to_workbook
            for t, row in latest.items():
                params = {k: row[k] for k in ("ticker", "company", "prospectus_date",
                          "lockup_days", "insider_pct", "ev_sales", "ev_ebitda",
                          "d_score", "modifier")}
                params.update(sec_source=row.get("sec_source", ""),
                              sponsor=row.get("sponsor", ""),
                              early_release=row.get("early_release", "No"))
                try:
                    file_bytes, s, ti, _ = add_name_to_workbook(file_bytes, params)
                    log.info(f"  Promoted {t} from bench (score {s})")
                except Exception as e:
                    log.error(f"  Promotion failed for {t}: {e}")
        except Exception as e:
            log.error(f"Promotion lookup failed: {e}")
    return file_bytes


def update_backtest_prices(file_bytes):
    """Write actual T-1/T+1/T+5/T+10 closes and % changes as VALUES into
    Historical Backtest for names whose lockup has passed. Uses the real
    trading calendar from yfinance (WORKDAY() misses market holidays, and
    STOCKHISTORY needs a live Excel 365 feed). Future dates stay blank."""
    from openpyxl import load_workbook as _lwb
    import datetime as _dt
    wb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
    hb = wb["Historical Backtest"]
    today = date.today()
    filled = 0
    for r in range(12, 200):
        t = hb.cell(row=r, column=2).value
        if not t:
            continue
        t = str(t).strip()
        if t.startswith(("NOTE", "SECTION")):
            continue
        d = hb.cell(row=r, column=4).value
        if not d:
            continue
        exp = d.date() if isinstance(d, _dt.datetime) else d
        if exp > today:
            continue
        try:
            import yfinance as yf
            hc = _get_history(t)
            if hc is not None and not hc.empty and \
               min(d2.date() for d2 in hc.index) <= exp - timedelta(days=3):
                h = hc
            else:
                time.sleep(0.4)
                h = yf.Ticker(t).history(start=exp - timedelta(days=15),
                                         end=exp + timedelta(days=25))
            if h is None or h.empty:
                continue
            closes = h["Close"]
            days = [d2.date() for d2 in closes.index]
            before = [i for i, d2 in enumerate(days) if d2 < exp]
            after  = [i for i, d2 in enumerate(days) if d2 > exp]
            if not before:
                continue
            p = {"G": float(closes.iloc[before[-1]])}
            for col, k in (("H", 1), ("I", 5), ("J", 10)):
                if len(after) >= k and days[after[k-1]] <= today:
                    p[col] = float(closes.iloc[after[k-1]])
            for col, v in p.items():
                cell = hb[f"{col}{r}"]
                cell.value = round(v, 2)
                cell.number_format = "$#,##0.00"
            for col, ref in (("K", "H"), ("L", "I"), ("M", "J")):
                cell = hb[f"{col}{r}"]
                if ref in p:
                    cell.value = round(p[ref] / p["G"] - 1, 4)
                    cell.number_format = "0.0%"
                else:
                    cell.value = None
            # date helpers: real trading dates as formatted values (no serials)
            helper = {"N": days[before[-1]]}
            for col, k in (("O", 1), ("P", 5), ("Q", 10)):
                if len(after) >= k:
                    helper[col] = days[after[k-1]]
            for col, dv in helper.items():
                cell = hb[f"{col}{r}"]
                cell.value = dv
                cell.number_format = "dd-mmm-yyyy"
            filled += 1
        except Exception as e:
            log.error(f"  Backtest price fill failed for {t}: {e}")
    out = io.BytesIO()
    wb.save(out)
    log.info(f"Backtest prices filled for {filled} expired names")
    return out.getvalue()

def sync_sources(file_bytes):
    """Ensure every Short Screen ticker has at least one Sources row (Lockup
    Date, Tier 1, from lockup_entries sec_source + EDGAR link). Automation-added
    names historically skipped Sources (lockup_engine gap)."""
    from openpyxl import load_workbook as _lwb
    from openpyxl.styles import Font
    wb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
    if "Sources" not in wb.sheetnames:
        return file_bytes
    ss, srcs = wb["Short Screen"], wb["Sources"]
    have = {str(srcs.cell(row=r, column=2).value).strip()
            for r in range(5, srcs.max_row + 1) if srcs.cell(row=r, column=2).value}
    screen = []
    for r in range(5, 300):
        t = ss.cell(row=r, column=3).value
        if not t:
            break
        screen.append((str(t).strip(), ss.cell(row=r, column=4).value))
    missing = [(t, c) for t, c in screen if t not in have]
    if not missing:
        return file_bytes
    meta = {}
    try:
        res = supabase.table("lockup_entries").select("ticker,sec_source")             .order("created_at", desc=True).execute()
        for row in (res.data or []):
            meta.setdefault(str(row["ticker"]).upper().strip(), row.get("sec_source") or "")
    except Exception as e:
        log.error(f"sec_source lookup failed: {e}")
    nr = srcs.max_row + 1
    for t, company in missing:
        sec = meta.get(t.upper(), "") or "SEC 424B4 (see EDGAR)"
        vals = {2: t, 3: company, 4: "Lockup Date", 5: sec, 6: "Tier 1",
                7: date.today().strftime("%b %Y"),
                8: "Auto-added by scanner; verify clause page in filing",
                9: f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&company={t}&type=424B4",
                10: "Current — Tier 1 SEC prospectus"}
        for c, v in vals.items():
            cell = srcs.cell(row=nr, column=c)
            cell.value = v
            cell.font = Font(name="Calibri", size=10)
        nr += 1
    out = io.BytesIO()
    wb.save(out)
    log.info(f"Sources rows added for {[t for t, _ in missing]}")
    return out.getvalue()

import re as _re2

_GREY_RGBS = {"FF808080", "FFA6A6A6", "FFBFBFBF", "FFD9D9D9", "FF7F7F7F", "FF595959"}

def _is_grey(color):
    """Grey detection across both encodings: literal ARGB and Excel theme
    colors (theme 0/1 with a tint = greyscale). Formatting bugs recur when
    only one encoding is handled."""
    try:
        if color.rgb and str(color.rgb).upper() in _GREY_RGBS:
            return True
        if color.rgb is None and color.theme in (0, 1) and abs(color.tint or 0) > 0.15:
            return True
    except Exception:
        pass
    return False

def finalize_scores(file_bytes):
    """Authoritative LAST pass each run: sanitize garbage momentum (nan price
    data must contribute F=0, not the +30 cap), recompute every score in
    Python, write Score/Tier as VALUES (SM O/P + SS L/M), and re-sort
    SS/SM/LV/HB descending with gate-FAIL names zeroed at the bottom.
    Runs after movement.py, so upstream formula/sort corruption cannot
    reach the delivered file."""
    from openpyxl import load_workbook as _lwb
    wb = _lwb(io.BytesIO(file_bytes), keep_vba=True)
    ss, sm = wb["Short Screen"], wb["Scoring Model"]
    lv = wb["Lockup Verification"] if "Lockup Verification" in wb.sheetnames else None
    hb = wb["Historical Backtest"] if "Historical Backtest" in wb.sheetnames else None

    def is_t(v):
        s = str(v or "").strip()
        return 1 <= len(s) <= 6 and s == s.upper() and " " not in s and any(c.isalpha() for c in s)

    n = 0
    scores = {}
    for r in range(5, 300):
        t = sm.cell(row=r, column=3).value
        if not is_t(t):
            break
        t = str(t).strip(); n += 1
        E  = float(sm.cell(row=r, column=5).value or 0)
        Fl = float(sm.cell(row=r, column=6).value or 0.01)
        I  = sm.cell(row=r, column=9).value
        J  = sm.cell(row=r, column=10).value
        L  = float(sm.cell(row=r, column=12).value or 0)
        M  = float(sm.cell(row=r, column=13).value or 0)
        Q  = float(sm.cell(row=r, column=17).value or 0)
        S  = sm.cell(row=r, column=19).value
        thesis = str(ss.cell(row=r, column=14).value or "")
        # movement.py writes F into the thesis line but not reliably into the
        # F column — parse the freshest momentum F and persist it to Q
        if "Momentum update:" in thesis and "nan%" not in thesis:
            m = _re2.findall(r"Sector \([^)]*\): F=([+-]?\d+)", thesis)
            if m:
                parsed = float(m[-1])
                if parsed != Q:
                    log.info(f"  Momentum F persisted {t}: Q {Q:+.0f} -> {parsed:+.0f} (from thesis)")
                    sm.cell(row=r, column=17).value = parsed
                    Q = parsed
        if "nan%" in thesis and Q != 0:
            log.info(f"  Momentum sanitized {t}: F {Q:+.0f} -> 0 (no price data)")
            sm.cell(row=r, column=17).value = 0
            Q = 0.0
            ss.cell(row=r, column=14).value = _re2.sub(
                r"\| Momentum update: 1W nan% / 1M nan% \| Sector \([^)]*\): F=[+-]?\d+",
                "| Momentum: no price data this run — F neutralized to 0", thesis)
        base, _t2 = calc_score(E, Fl, I if I is not None else "NM",
                               J if J is not None else "NM", L, M)
        raw = base + Q
        score = int(excel_round(max(0, min(100, raw)), 0))
        failed = S in ("FAIL", "ILLIQUID")
        eff = 0 if failed else score
        tier = "ILLIQUID" if failed else (
            "High" if eff >= 75 else "Medium" if eff >= 50 else
            "Low" if eff >= 25 else "Minimal")
        scores[t] = (eff, tier, failed, raw)

    for r in range(5, 5 + n):
        t = str(sm.cell(row=r, column=3).value).strip()
        eff, tier, _f, _raw = scores[t]
        sm.cell(row=r, column=15).value = eff
        sm.cell(row=r, column=16).value = tier
        ss.cell(row=r, column=12).value = eff
        ss.cell(row=r, column=13).value = tier

    order = sorted(scores, key=lambda t: (not scores[t][2], scores[t][0], scores[t][3]),
                   reverse=True)

    def _rw(f, r):
        return _re2.sub(r"([A-Z]{1,2})\d+(?![0-9])",
                        lambda m: f"{m.group(1)}{r}", f)

    def reorder(ws, tc, start, rank=False):
        rows = {}
        r = start
        while is_t(ws.cell(row=r, column=tc).value):
            rows[str(ws.cell(row=r, column=tc).value).strip()] = \
                [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            r += 1
        if set(rows) != set(order):
            log.error(f"  reorder skipped on {ws.title}: ticker set mismatch")
            return
        for i, t in enumerate(order):
            rr = start + i
            for c, v in enumerate(rows[t], start=1):
                ws.cell(row=rr, column=c).value = v
            if rank:
                ws.cell(row=rr, column=2).value = i + 1
            for c in range(2, ws.max_column + 1):
                f = ws.cell(row=rr, column=c).value
                if isinstance(f, str) and f.startswith("="):
                    ws.cell(row=rr, column=c).value = _rw(f, rr)

    reorder(ss, 3, 5, rank=True)
    reorder(sm, 3, 5, rank=True)
    if lv is not None:
        reorder(lv, 2, 5)
    if hb is not None:
        reorder(hb, 2, 12)

    # Re-paint score/tier cells BY TIER (the template's banding was positional,
    # so re-sorting left colors stuck to old row positions — grey mid-table)
    from openpyxl.styles import PatternFill, Font as _Font
    PALETTE = {  # fill, font color, bold
        "High":     ("FF1F3864", "FFFFFFFF", True),
        "Medium":   ("FFFF8C00", "FFFFFFFF", True),
        "Low":      ("FFFFC000", "FF000000", False),
        "Minimal":  ("FF92D050", "FF000000", False),
        "ILLIQUID": ("FFD9D9D9", "FF808080", False),
    }
    _no_fill = PatternFill(fill_type=None)
    for i, t in enumerate(order):
        rr = 5 + i
        tier = scores[t][1]
        fill_hex, font_hex, bold = PALETTE.get(tier, PALETTE["Minimal"])
        fill = PatternFill(fill_type="solid", start_color=fill_hex, end_color=fill_hex)
        for ws, tier_cols in ((ss, (12, 13)), (sm, (15, 16))):
            # clear the positional grey row band across the whole data row,
            # then paint only the score/tier cells by tier
            for c in range(2, ws.max_column + 1):
                if c in tier_cols:
                    continue
                cell = ws.cell(row=rr, column=c)
                cur = cell.fill
                if cur is not None and cur.fill_type == "solid":
                    cell.fill = _no_fill
                f = cell.font
                if f is not None and f.color is not None and not f.italic \
                        and _is_grey(f.color):
                    # grey-band text back to black (leave gray-italic footnotes
                    # alone). _is_grey also catches Excel theme-encoded greys,
                    # which appear the first time the file is saved in Excel.
                    cell.font = _Font(name=f.name or "Calibri", size=f.size or 11,
                                      bold=f.b, italic=False, color="FF000000")
            for c in tier_cols:
                cell = ws.cell(row=rr, column=c)
                cell.fill = fill
                old_f = cell.font
                cell.font = _Font(name=old_f.name or "Calibri", size=old_f.size or 11,
                                  bold=bold, italic=(tier == "ILLIQUID"),
                                  color=font_hex)

    out = io.BytesIO()
    wb.save(out)
    log.info(f"Scores finalized as values; top: {[(t, scores[t][0]) for t in order[:3]]}")
    return out.getvalue()

def run():
    log.info("=== Starting lockup scan ===")

    filings = fetch_recent_filings(days_back=7)
    if not filings:
        log.info("No filings found.")
        return

    try:
        file_bytes = download_workbook()
    except Exception as e:
        log.error(f"Workbook download failed: {e}")
        return

    # Pre-filter obvious non-IPO lockups before hitting Claude API
    SKIP_KEYWORDS = [
        'acquisition corp', 'acquisition inc', 'holdings xi', 'holdings x ',
        'holdings ix', 'holdings viii', 'holdings vii', 'holdings vi',
        'ventures acquisition', 'equity partners', 'capital solutions',
        'spac', 'blank check', 'wilco', 'tianci',
    ]

    def is_likely_ipo(company_name):
        name_lower = company_name.lower()
        for kw in SKIP_KEYWORDS:
            if kw in name_lower:
                log.info(f"  Pre-filter skipped: {company_name} (matched: {kw})")
                return False
        return True

    filtered_filings = [f for f in filings if is_likely_ipo(f["entity_name"])]
    log.info(f"After pre-filter: {len(filtered_filings)} of {len(filings)} filings remain")

    MAX_CLAUDE_CALLS = 3  # hard cap to control API cost
    candidates = []
    seen_tickers = set()

    # Robust dedup: seed with tickers ALREADY in the workbook (source of truth),
    # not the lockup_entries table. The table can contain names that never landed
    # in the workbook (failed add/upload) — those must remain eligible so the
    # reconcile step can backfill them.
    try:
        seen_tickers |= get_workbook_tickers(file_bytes)
        log.info(f"Dedup seeded with {len(seen_tickers)} existing workbook tickers")
    except Exception as e:
        log.error(f"Could not seed dedup from workbook: {e}")

    # Backfill any table entries missing from the workbook (uses stored research,
    # zero Claude calls). Fixes orphans like BLSM/ATTO/APMD/CISS/AGCC.
    try:
        file_bytes, backfilled = reconcile_table_to_workbook(file_bytes)
        if backfilled:
            seen_tickers |= get_workbook_tickers(file_bytes)
            log.info(f"Reconcile: {backfilled} names backfilled (in-memory)")
    except Exception as e:
        log.error(f"Reconcile failed: {e}")

    claude_calls = 0
    for f in filtered_filings[:10]:
        if claude_calls >= MAX_CLAUDE_CALLS:
            log.info(f"Claude call cap ({MAX_CLAUDE_CALLS}) reached — stopping research")
            break
        log.info(f"Researching: {f['entity_name']}")
        data = research_ticker(f["entity_name"], f["file_date"])
        claude_calls += 1
        if not data: continue
        if data.get("skip"):
            log.info(f"  Skipped: {data.get('skip_reason')}")
            continue
        ticker = data.get("ticker","").upper().strip()
        if not ticker or ticker in ("N/A","TBD",""): continue
        if ticker in seen_tickers:
            log.info(f"  {ticker} duplicate this run, skipping")
            continue
        seen_tickers.add(ticker)
        # NOTE: no lockup_entries check here — the table is NOT the source of
        # truth for what's in the workbook. seen_tickers (workbook-seeded)
        # already handles dedup; a table row without a workbook row is an
        # orphan handled by reconcile_table_to_workbook().
        ok, reason = qualifies(ticker)
        if not ok:
            log.info(f"  {ticker} fails universe rules ({reason}) — benched, not added")
            continue
        insider = float(data.get("insider_pct") or 0)
        score, tier = calc_score(insider, round(100-insider,2),
                                  data.get("ev_sales","NM"), data.get("ev_ebitda","NM"),
                                  float(data.get("d_score") or 0), float(data.get("modifier") or 0))
        data["_score"] = score; data["_tier"] = tier
        candidates.append(data)
        time.sleep(1)

    candidates.sort(key=lambda x: x["_score"], reverse=True)
    candidates = candidates[:MAX_PER_RUN]
    log.info(f"Adding {len(candidates)} candidates")

    added = 0
    for data in candidates:
        try:
            from lockup_engine import add_name_to_workbook
            file_bytes, score, tier, days_out = add_name_to_workbook(file_bytes, data)
            log_entry(data, score, tier)
            added += 1
            log.info(f"Added {data['ticker']}")
        except Exception as e:
            log.error(f"Failed to add {data['ticker']}: {e}")

    # (no mid-run upload — single upload happens at end of pipeline)

    # Update momentum (price + Form 4) for all tickers — free, no Claude API.
    # CRITICAL: the whole pipeline now runs on ONE in-memory copy with a single
    # upload at the very end. Supabase Storage GETs are CDN-cached, so a
    # download issued milliseconds after an upload returns the STALE object —
    # the old download->upload round-trip between steps was erasing momentum's
    # writes on every run (PUT 12:59:36.176 -> stale GET 12:59:36.254).
    log.info("Updating momentum in thesis cells...")
    latest = file_bytes  # continue from the in-memory copy (includes any adds)
    try:
        from movement import update_momentum
        latest = update_momentum(latest)
        log.info("Momentum update complete.")
    except Exception as e:
        log.error(f"Momentum update failed: {e}")

    # Keep Historical Backtest in lockstep with the Short Screen (self-row
    # formulas, contiguous rows, no merged-cell corruption).
    try:
        latest = enforce_universe_rules(latest)
        latest = sync_sources(latest)
        latest = sync_historical_backtest(latest)
        latest = update_backtest_prices(latest)
        latest = resolve_liquidity_gate(latest)
        latest = finalize_scores(latest)
        upload_workbook(latest)
        log.info("Single end-of-run upload complete.")
    except Exception as e:
        log.error(f"Final pipeline failed: {e}")
        # salvage: upload whatever stage succeeded so momentum isn't lost
        try:
            upload_workbook(latest)
            log.info("Salvage upload of last good stage complete.")
        except Exception as e2:
            log.error(f"Salvage upload failed: {e2}")

    log.info(f"=== Done. {added} names added. ===")
