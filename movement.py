"""
Momentum updater — appends live 1W/1M price + Form 4 activity
to the thesis cell (col N) for each ticker on Short Screen.
Zero Anthropic API usage. Free data only: Yahoo Finance + SEC EDGAR.
"""
import io, time, logging, re, requests
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

log = logging.getLogger(__name__)

HEADERS = {"User-Agent": "Brandon Ross brandonr1010@gmail.com"}
MOMENTUM_TAG = "| Momentum update:"  # tag so we can find/replace the appended line

def get_cik(ticker):
    """Look up CIK from SEC EDGAR company tickers JSON."""
    try:
        r = requests.get(
            "https://data.sec.gov/submissions/company_tickers.json",
            headers=HEADERS, timeout=15
        )
        if r.status_code != 200:
            return None
        data = r.json()
        for k, v in data.items():
            if v.get('ticker', '').upper() == ticker.upper():
                return str(v['cik_str']).zfill(10)
        return None
    except:
        return None

def get_form4_activity(cik, days_back=30):
    """
    Check recent Form 4 filings for a CIK.
    Returns a short string like "2 insider sales" or "No Form 4s filed"
    """
    try:
        url = f"https://data.sec.gov/submissions/CIK{cik}.json"
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        recent = data.get('filings', {}).get('recent', {})
        forms = recent.get('form', [])
        dates = recent.get('filedAt', recent.get('filingDate', []))
        cutoff = (date.today() - timedelta(days=days_back)).isoformat()
        form4s = [dates[i] for i in range(len(forms)) if forms[i] == '4' and dates[i] >= cutoff]
        if not form4s:
            return f"No Form 4s past {days_back}d"
        return f"{len(form4s)} Form 4 filing{'s' if len(form4s)>1 else ''} past {days_back}d"
    except:
        return None

def get_price_momentum(ticker):
    """Returns (1W_pct, 1M_pct) via Yahoo Finance."""
    try:
        import yfinance as yf
        hist = yf.Ticker(ticker).history(period="35d")
        if hist.empty or len(hist) < 2:
            return None, None
        cur = hist['Close'].iloc[-1]
        w1 = round((cur - hist['Close'].iloc[-6]) / hist['Close'].iloc[-6] * 100, 1) if len(hist) >= 6 else None
        m1 = round((cur - hist['Close'].iloc[-22]) / hist['Close'].iloc[-22] * 100, 1) if len(hist) >= 22 else None
        return w1, m1
    except:
        return None, None

def fmt_pct(val):
    if val is None: return "N/A"
    return f"+{val}%" if val >= 0 else f"{val}%"

def strip_old_momentum(thesis):
    """Remove previously appended momentum line."""
    if not thesis:
        return thesis
    idx = thesis.find(MOMENTUM_TAG)
    if idx != -1:
        return thesis[:idx].rstrip()
    return thesis

def update_momentum(file_bytes, newsapi_key=None):
    """
    For each ticker in Short Screen:
    1. Pull 1W/1M price from Yahoo Finance
    2. Pull recent Form 4 count from EDGAR
    3. Pull sector ETF momentum + news sentiment (F score)
    4. Append/replace momentum line in thesis cell (col N)
    Returns updated file bytes.
    """
    import os
    if not newsapi_key:
        newsapi_key = os.environ.get("NEWSAPI_KEY")

    wb = load_workbook(io.BytesIO(file_bytes), keep_vba=True)
    wsSS = wb["Short Screen"]

    # Get all tickers first
    tickers = []
    for r in range(5, 200):
        t = wsSS.cell(row=r, column=3).value
        if not t: break
        tickers.append(t)

    # Cache CIK lookup
    cik_map = {}
    try:
        r = requests.get(
            "https://data.sec.gov/submissions/company_tickers.json",
            headers=HEADERS, timeout=15
        )
        if r.status_code == 200:
            data = r.json()
            for k, v in data.items():
                cik_map[v.get('ticker','').upper()] = str(v['cik_str']).zfill(10)
            log.info(f"CIK map loaded: {len(cik_map)} tickers")
        time.sleep(0.3)
    except Exception as e:
        log.warning(f"CIK batch load failed: {e}")

    # Get sector context (ETF + news per sector, cached). Stock momentum added per-ticker.
    sector_context = {}
    try:
        from sector import get_sector_context
        sector_context = get_sector_context(tickers, newsapi_key)
        log.info(f"Sector context computed for {len(sector_context)} sectors")
    except Exception as e:
        log.error(f"Sector context FAILED: {type(e).__name__}: {e}")
        import traceback
        log.error(traceback.format_exc())
        sector_context = {}

    updated = 0
    for row in range(5, 200):
        ticker = wsSS.cell(row=row, column=3).value
        if not ticker: break

        log.info(f"  Updating momentum for {ticker}")

        # Price momentum
        w1, m1 = get_price_momentum(ticker)
        time.sleep(0.3)

        # Form 4 activity
        f4_str = None
        cik = cik_map.get(ticker.upper())
        if cik:
            f4_str = get_form4_activity(cik, days_back=30)
            time.sleep(0.35)

        # Compute F using stock momentum (w1/m1) + cached sector context
        try:
            from sector import score_ticker
            f_score, sector, _bd = score_ticker(ticker, w1, m1, sector_context)
        except Exception as e:
            log.warning(f"  {ticker}: score_ticker failed: {e}")
            f_score, sector = 0, "Unknown"

        # Build momentum line
        price_str = f"1W {fmt_pct(w1)} / 1M {fmt_pct(m1)}"
        f4_part = f" | {f4_str}" if f4_str else ""
        sector_part = f" | Sector ({sector}): F={'+' if f_score>=0 else ''}{f_score}"
        momentum_line = f"{MOMENTUM_TAG} {price_str}{f4_part}{sector_part}"

        # Write F score to SM col 17 (Q) — feeds into Raw formula
        sm_f_cell = wb["Scoring Model"].cell(row=row, column=17)
        if not isinstance(sm_f_cell, MergedCell):
            sm_f_cell.value = f_score
            sm_f_cell.number_format = '+0;-0;0'
            from openpyxl.styles import Alignment as Aln
            sm_f_cell.alignment = Aln(horizontal='center')

        # Update thesis cell
        thesis_cell = wsSS.cell(row=row, column=14)
        if isinstance(thesis_cell, MergedCell):
            continue
        old_thesis = str(thesis_cell.value) if thesis_cell.value else ""
        clean_thesis = strip_old_momentum(old_thesis)
        thesis_cell.value = f"{clean_thesis}  {momentum_line}" if clean_thesis else momentum_line
        updated += 1
        log.info(f"    {ticker}: {momentum_line}")

    log.info(f"Momentum updated for {updated} tickers — re-sorting by updated scores...")

    # Read prior ranks BEFORE re-sort (from hidden state sheet)
    prior_state = _read_prior_state(wb)

    # Re-sort SS and SM after F scores are written
    try:
        _resort_after_f_update(wb)
        log.info("Re-sort complete.")
    except Exception as e:
        log.error(f"Re-sort failed: {e}")

    # Capture current ranks after re-sort (row position = rank)
    current_state = []
    wsSS_after = wb["Short Screen"]
    for r in range(5, 200):
        tk = wsSS_after.cell(row=r, column=3).value
        if not tk: break
        current_state.append((tk, r - 4))  # (ticker, rank)

    # Build/refresh Daily Movers tab (rank-based: top 5 climbers, top 5 fallers)
    try:
        _build_daily_movers(wb, current_state, prior_state)
        log.info("Daily Movers tab updated.")
    except Exception as e:
        log.error(f"Daily Movers build failed: {e}")
        import traceback
        log.error(traceback.format_exc())

    # Persist current ranks for next scan's comparison
    try:
        _write_current_state(wb, current_state)
    except Exception as e:
        log.error(f"State write failed: {e}")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _resort_after_f_update(wb):
    """Re-sort Short Screen and Scoring Model by updated scores after F is written."""
    from copy import copy as xcopy
    from openpyxl.styles import PatternFill
    from openpyxl.cell.cell import MergedCell as MC

    wsSS = wb["Short Screen"]
    wsSM = wb["Scoring Model"]

    TIER_COLORS = {"High":"FF1F3864","Medium":"FFFF8C00","Low":"FFFFC000","Minimal":"FF92D050"}
    TIER_ROW_BG = {"High":"FF1F3864","Medium":"FFFFEB9C","Low":"FFFFFFFF","Minimal":"FFFFFFFF"}
    FONT_COLORS = {"High":"FFFFFFFF","Medium":"FF006400","Low":"FF006400","Minimal":"FF006400"}

    def get_score_from_inputs(r):
        """Compute score directly from SM inputs including F."""
        import math
        def er(val, dec=0):
            f = 10**dec; return math.floor(val*f+0.5)/f
        try:
            ins = float(wsSM.cell(row=r,column=5).value or 0)
            flt = float(wsSM.cell(row=r,column=6).value or 0)
            ev_s = wsSM.cell(row=r,column=9).value
            ev_e = wsSM.cell(row=r,column=10).value
            d = float(wsSM.cell(row=r,column=12).value or 0)
            e = float(wsSM.cell(row=r,column=13).value or 0)
            f = float(wsSM.cell(row=r,column=17).value or 0)
            A = er(min(ins/max(flt,0.01),5)/5*30,1)
            B = er(min(ins/100,1)*25,1)
            # Valuation Risk (0-20): no earnings = max (20). Must match calc_score.
            def _nm(x):
                try:
                    float(x); return False
                except (ValueError, TypeError):
                    return True
            if _nm(ev_s) and _nm(ev_e):
                cS, cE = 20.0, 0.0
            else:
                cS = 10.0 if _nm(ev_s) else min(float(ev_s)/5*9, 9)
                cE = 10.0 if _nm(ev_e) else min(max(float(ev_e)-5,0)/25*9, 9)
            C = er(cS+cE,1)
            F = er(max(-30,min(30,f)),1)
            raw = er(A+B+C+d+e+F,1)
            score = int(er(max(0,min(100,raw)),0))
            tier = "High" if score>=75 else "Medium" if score>=50 else "Low" if score>=25 else "Minimal"
            return score, tier
        except: return 0, "Minimal"

    # Collect all rows
    rows_ss = []; rows_sm = []
    last_row = 5
    while wsSS.cell(row=last_row,column=3).value: last_row+=1
    last_row -= 1

    for r in range(5, last_row+1):
        ss_data={c:wsSS.cell(row=r,column=c).value for c in range(2,15)}
        ss_fmt={c:wsSS.cell(row=r,column=c).number_format for c in range(2,15)}
        ss_fill={c:xcopy(wsSS.cell(row=r,column=c).fill) for c in range(2,15)}
        ss_font={c:xcopy(wsSS.cell(row=r,column=c).font) for c in range(2,15)}
        score,tier=get_score_from_inputs(r)
        rows_ss.append({'data':ss_data,'fmt':ss_fmt,'fill':ss_fill,'font':ss_font,'score':score,'tier':tier,'ticker':ss_data.get(3)})

        sm_data={c:(None if isinstance(wsSM.cell(row=r,column=c),MC) else wsSM.cell(row=r,column=c).value) for c in range(2,18)}
        sm_fmt={c:(wsSM.cell(row=r,column=c).number_format if not isinstance(wsSM.cell(row=r,column=c),MC) else '') for c in range(2,18)}
        sm_fill={c:(xcopy(wsSM.cell(row=r,column=c).fill) if not isinstance(wsSM.cell(row=r,column=c),MC) else None) for c in range(2,18)}
        sm_font={c:(xcopy(wsSM.cell(row=r,column=c).font) if not isinstance(wsSM.cell(row=r,column=c),MC) else None) for c in range(2,18)}
        rows_sm.append({'data':sm_data,'fmt':sm_fmt,'fill':sm_fill,'font':sm_font,'ticker':sm_data.get(3)})

    rows_ss.sort(key=lambda x:x['score'],reverse=True)
    ticker_order=[r['ticker'] for r in rows_ss]
    sm_by_ticker={r['ticker']:r for r in rows_sm}
    rows_sm_sorted=[sm_by_ticker[t] for t in ticker_order if t in sm_by_ticker]

    from openpyxl.styles import Font as Fnt
    for idx,row in enumerate(rows_ss):
        r=5+idx
        for c in range(2,15):
            cell=wsSS.cell(row=r,column=c)
            cell.value=row['data'][c]; cell.number_format=row['fmt'][c]
            cell.fill=row['fill'][c]; cell.font=row['font'][c]
        wsSS.cell(row=r,column=2).value=f"=ROW()-4"
        wsSS.cell(row=r,column=6).value=f"=E{r}-TODAY()"
        wsSS.cell(row=r,column=12).value=f"='Scoring Model'!O{r}"
        wsSS.cell(row=r,column=13).value=f"='Scoring Model'!P{r}"
        t=row['tier']
        sc=TIER_COLORS[t]; bg=TIER_ROW_BG[t]; fc=FONT_COLORS[t]
        for col in [2,3,4,5,6,7,8,9,10,11,14]:
            c2=wsSS.cell(row=r,column=col)
            if not isinstance(c2,MC): c2.fill=PatternFill("solid",start_color=bg,end_color=bg)
        for col in [12,13]:
            c2=wsSS.cell(row=r,column=col)
            if not isinstance(c2,MC):
                c2.fill=PatternFill("solid",start_color=sc,end_color=sc)
                c2.font=Fnt(bold=True,color=fc,name="Calibri",size=10)

    for idx,row in enumerate(rows_sm_sorted):
        r=5+idx
        for c in range(2,18):
            cell=wsSM.cell(row=r,column=c)
            if isinstance(cell,MC): continue
            cell.value=row['data'].get(c)
            if row['fmt'].get(c): cell.number_format=row['fmt'][c]
            if row['fill'].get(c): cell.fill=row['fill'][c]
            if row['font'].get(c): cell.font=row['font'][c]
        wsSM.cell(row=r,column=2).value=f"=ROW()-4"
        wsSM.cell(row=r,column=7).value=f"=IFERROR(ROUND(MIN(E{r}/MAX(F{r},0.01),5)/5*30,1),0)"
        wsSM.cell(row=r,column=8).value=f"=ROUND(MIN(E{r}/100,1)*25,1)"
        wsSM.cell(row=r,column=11).value=f'=ROUND(IF(AND(ISERROR(VALUE(I{r})),ISERROR(VALUE(J{r}))),20,IF(ISERROR(VALUE(I{r})),10,MIN(VALUE(I{r})/5*9,9))+IF(ISERROR(VALUE(J{r})),10,MIN(MAX(VALUE(J{r})-5,0)/25*9,9))),1)'
        wsSM.cell(row=r,column=14).value=f"=ROUND(G{r}+H{r}+K{r}+L{r}+M{r}+Q{r},1)"
        wsSM.cell(row=r,column=15).value=f"=MAX(0,MIN(100,ROUND(N{r},0)))"
        wsSM.cell(row=r,column=16).value=f'=IF(O{r}>=75,"High",IF(O{r}>=50,"Medium",IF(O{r}>=25,"Low","Minimal")))'
        s,t=get_score_from_inputs(r)
        t_str="High" if s>=75 else "Medium" if s>=50 else "Low" if s>=25 else "Minimal"
        sc=TIER_COLORS[t_str]; bg=TIER_ROW_BG[t_str]; fc=FONT_COLORS[t_str]
        from openpyxl.styles import Font as Fnt2
        for col in range(2,15):
            c2=wsSM.cell(row=r,column=col)
            if not isinstance(c2,MC): c2.fill=PatternFill("solid",start_color=bg,end_color=bg)
        for col in [15,16]:
            c2=wsSM.cell(row=r,column=col)
            if not isinstance(c2,MC):
                c2.fill=PatternFill("solid",start_color=sc,end_color=sc)
                c2.font=Fnt2(bold=True,color=fc,name="Calibri",size=10)

    # Ensure buffer row
    last_data=5+len(rows_sm_sorted)-1
    buf=last_data+1
    buf_cell=wsSM.cell(row=buf,column=3)
    if not isinstance(buf_cell,MC) and buf_cell.value and not str(buf_cell.value).startswith('='):
        wsSM.insert_rows(buf)


# ─── DAILY MOVERS (rank-based) ───────────────────────────────────
STATE_SHEET = "_movers_state"

def _read_prior_state(wb):
    """Read prior {ticker: rank} from hidden state sheet. Empty dict on first run."""
    if STATE_SHEET not in wb.sheetnames:
        return {}
    ws = wb[STATE_SHEET]
    prior = {}
    for r in range(2, 500):
        tk = ws.cell(row=r, column=1).value
        if not tk:
            break
        try:
            prior[tk] = int(ws.cell(row=r, column=3).value)
        except (ValueError, TypeError):
            continue
    return prior

def _write_current_state(wb, current_state):
    """Overwrite hidden state sheet with current [(ticker, rank), ...]."""
    from openpyxl.cell.cell import MergedCell as MC
    if STATE_SHEET in wb.sheetnames:
        ws = wb[STATE_SHEET]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MC):
                    cell.value = None
    else:
        ws = wb.create_sheet(STATE_SHEET)
    ws.cell(row=1, column=1).value = "ticker"
    ws.cell(row=1, column=2).value = "score"
    ws.cell(row=1, column=3).value = "rank"
    for i, (tk, rank) in enumerate(current_state):
        ws.cell(row=2+i, column=1).value = tk
        ws.cell(row=2+i, column=3).value = rank
    ws.sheet_state = "hidden"
    return ws

def _build_daily_movers(wb, current_state, prior_state):
    """
    Daily Movers tab: top 5 names that climbed the most ranks toward #1 (toward SHORT)
    and top 5 that fell the most ranks toward the bottom (toward LONG), since last scan.
    rank_change = prior_rank - current_rank  (positive = climbed toward short).
    Purely additive — touches only this tab. Rebuilt fresh each scan.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell.cell import MergedCell as MC
    from datetime import datetime

    NAVY="FF1F3864"; WHITE="FFFFFFFF"; LIGHT="FFF0F4FA"
    GREEN="FF006400"; RED="FFC00000"; GRAY="FF808080"
    UP_BG="FFE2EFDA"; DOWN_BG="FFFCE4E4"

    SHEET="Daily Movers"
    if SHEET in wb.sheetnames:
        ws=wb[SHEET]
        # Unmerge any existing merged ranges first (stale merges break re-runs)
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell,MC):
                    cell.value=None
                    cell.fill=PatternFill(fill_type=None)
    else:
        ws=wb.create_sheet(SHEET)
        try:
            target=wb.sheetnames.index("Short Screen")+1
            wb.move_sheet(SHEET, offset=target-wb.sheetnames.index(SHEET))
        except Exception:
            pass

    # Company lookup from Short Screen
    wsSS=wb["Short Screen"]
    company={}
    for r in range(5,200):
        tk=wsSS.cell(row=r,column=3).value
        if not tk: break
        company[tk]=wsSS.cell(row=r,column=4).value or ""

    # Compute rank changes (only names present in both prior and current)
    movers=[]
    for tk, cur_rank in current_state:
        if tk in prior_state:
            change = prior_state[tk] - cur_rank  # + = climbed toward #1
            if change != 0:
                movers.append({"ticker":tk,"company":company.get(tk,""),
                               "prior":prior_state[tk],"current":cur_rank,"change":change})

    climbers=sorted([m for m in movers if m["change"]>0], key=lambda m:m["change"], reverse=True)[:5]
    fallers =sorted([m for m in movers if m["change"]<0], key=lambda m:m["change"])[:5]

    thin=Side(style="thin",color="FFD0D0D0")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)

    # Title
    ws.merge_cells("B2:G2")
    t=ws["B2"]
    t.value=f"DAILY MOVERS — Biggest Rank Moves (updated {datetime.now().strftime('%m/%d/%Y')})"
    t.fill=PatternFill("solid",start_color=NAVY,end_color=NAVY)
    t.font=Font(bold=True,color=WHITE,size=11,name="Calibri")
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=22

    def section(start_row, label, label_bg, rows_data, empty_msg):
        # Section banner
        ws.merge_cells(start_row=start_row,start_column=2,end_row=start_row,end_column=7)
        b=ws.cell(row=start_row,column=2)
        b.value=label
        b.fill=PatternFill("solid",start_color=label_bg,end_color=label_bg)
        b.font=Font(bold=True,color=WHITE,size=10,name="Calibri")
        b.alignment=Alignment(horizontal="center",vertical="center")
        # Headers
        hdrs=["Ticker","Company","Prior Rank","Today Rank","Ranks Moved"]
        cols=[2,3,4,5,6,7]
        hr=start_row+1
        for col,h in zip([2,3,4,5,6],hdrs):
            c=ws.cell(row=hr,column=col)
            c.value=h
            c.fill=PatternFill("solid",start_color=NAVY,end_color=NAVY)
            c.font=Font(bold=True,color=WHITE,size=9,name="Calibri")
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=border
        # Data
        if not rows_data:
            ws.merge_cells(start_row=hr+1,start_column=2,end_row=hr+1,end_column=6)
            c=ws.cell(row=hr+1,column=2)
            c.value=empty_msg
            c.font=Font(italic=True,color=GRAY,size=9)
            c.alignment=Alignment(horizontal="center")
            return hr+2
        for i,m in enumerate(rows_data):
            r=hr+1+i
            up=m["change"]>0
            bg=UP_BG if up else DOWN_BG
            arrow="▲" if up else "▼"
            mc=GREEN if up else RED
            cells=[
                (2,m["ticker"],"center",Font(bold=True,color="FF000000",size=10,name="Calibri")),
                (3,m["company"],"left",Font(color="FF000000",size=9,name="Calibri")),
                (4,m["prior"],"center",Font(color="FF000000",size=9,name="Calibri")),
                (5,m["current"],"center",Font(color="FF000000",size=9,name="Calibri")),
                (6,f"{arrow} {abs(m['change'])}","center",Font(bold=True,color=mc,size=10,name="Calibri")),
            ]
            for col,val,align,font in cells:
                c=ws.cell(row=r,column=col)
                c.value=val
                c.fill=PatternFill("solid",start_color=bg,end_color=bg)
                c.alignment=Alignment(horizontal=align,vertical="center")
                c.border=border
                c.font=font
        return hr+1+len(rows_data)+1

    next_row = section(4, "▲ TOP 5 CLIMBERS — Moving Toward SHORT (top of screen)",
                        "FF1F3864", climbers, "No upward rank moves this scan.")
    next_row += 1
    section(next_row, "▼ TOP 5 FALLERS — Moving Toward LONG (bottom of screen)",
            "FFC00000", fallers, "No downward rank moves this scan.")

    widths={"A":2,"B":8,"C":30,"D":11,"E":11,"F":13,"G":4}
    for col,w in widths.items():
        ws.column_dimensions[col].width=w
    ws.sheet_view.showGridLines=False
    return ws
